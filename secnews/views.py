from django.shortcuts import render
from django.http import FileResponse,HttpResponse
import requests as req
import urllib.request
from bs4 import BeautifulSoup as bs4
import io
import csv
import openpyxl as px

list = []#ニュースの情報を取得

#ルールベースで分類した結果を格納するもの
incident_list = []#インシデントに関するニュース
defense_list = []#セキュリティ対策に関するニュース
other_list = []#コラムや調査などといったその他のニュース

overseas_list =[]

#スクレイピングしたものを欠損値処理をした上で格納
def list_appends(sites,topic,topurl):
    
    for element in topic.find_all("a"):      
        
        if element.text == "\n\n\n\n":
            continue
        else:
            if len(element.text) != 0:
                
                check_url = element.get("href")
                
                if check_url.startswith('h'):
                    list.append([sites,element.text,element.get("href")])
                else:
                    list.append([sites,element.text,topurl+element.get("href")])
                
        
        for i in range(len(list)-1):
            if list[i] == element.text:
                del list[-1]

#スクレイピングしたものを欠損値処理をした上で格納。
#海外ニュースの場合はこちらを適用
def overseas_list_appends(sites,topic,topurl):
    
    for element in topic.find_all("a"):      
        
        if element.text == "\n\n\n\n":
            continue
        else:
            if len(element.text) != 0:
                
                check_url = element.get("href")
                title = element.text
                
                if title.startswith('\n\n'):
                    title= title.strip('\n\n')
                
                if title.endswith('new'):
                    title = title.rstrip('new')
                
                if check_url.startswith('h'):
                    overseas_list.append([sites,title,element.get("href")])
                else:
                    overseas_list.append([sites,title,topurl+element.get("href")])
                
        
        for i in range(len(overseas_list)-1):
            if overseas_list[i] == element.text:
                del list[-1]

#スクレイピングしたものを欠損値処理をした上で格納。
#ただし、2つのタグを別々にスクレイピングする必要がある場合に使用する。
def list_appends_separation(tag_topic,tag_url,sites,topic,topurl):
    
    process_list_title = []
    process_list_url = []   
    
    for element in topic.find_all(tag_topic):
        
        if element.text == "\n\n\n\n":
            continue
        else:
            if len(element.text) != 0:
                process_list_title.append(element.text)

    for element in topic.find_all(tag_url):
        
        if element.text == "\n\n\n\n":
            continue
        else:
            if len(element.text) != 0:
                process_list_url.append(element.get("href"))
    
    for i in range(len(process_list_url)):
        check_url = process_list_url[i]
        if check_url.startswith('h'):
            list.append([sites,process_list_title[i],process_list_url[i]])
        else:
            list.append([sites,process_list_title[i],topurl+process_list_url[i]])
    
    process_list_title.clear()
    process_list_url.clear()

#スクレイピングしたものを欠損値処理をした上で格納。
#ただし、2つのタグを別々にスクレイピングする必要がある場合に使用する。
#海外ニュースの場合に適用
def overseas_list_appends_separation(tag_topic,tag_url,sites,topic,topurl):
    
    process_list_title = []
    process_list_url = []   
    
    for element in topic.find_all(tag_topic):
        
        if element.text == "\n\n\n\n":
            continue
        else:
            if len(element.text) != 0:
                process_list_title.append(element.text)

    for element in topic.find_all(tag_url):
        
        if element.text == "\n\n\n\n":
            continue
        else:
            if len(element.text) != 0:
                process_list_url.append(element.get("href"))
    
    for i in range(len(process_list_url)):
        check_url = process_list_url[i]
        if check_url.startswith('h'):
            overseas_list.append([sites,process_list_title[i],process_list_url[i]])
        else:
            overseas_list.append([sites,process_list_title[i],topurl+process_list_url[i]])
    
    process_list_title.clear()
    process_list_url.clear()

#ルールベースでの分類
def list_classification(sites,title,url):
    flag = False
    
    #辞書
    word_advance = ["脆弱性対策","注意喚起"]
    word_incident = ["流失","流出","脆弱性","サイバー攻撃","詐欺","攻撃","被害",
                     "不正","処分","ミス","被害","所在不明","悪用","感染","不備",
                    "だます","なりすまし","別人","紛失","閉鎖","利用不可","置き忘れ",
                     "注意","警告","まずい","中止",]
    word_defense = ["診断","診断機能","ゼロトラスト","対策","訓練","体験","教育","演習","打ち手","封じ込め","評価"]

    #分類処理
  
    for i in range(len(word_advance)):
    
        if word_advance[i] in title:
            incident_list.append([sites,title,url])
            flag = True
            break
        else:
            pass
    
    #分類処理
    
    if flag == False:
        
        for i in range(len(word_defense)):
    
            if word_defense[i] in title:
                defense_list.append([sites,title,url])
                flag = True
                break
            else:
                pass
            
    if flag == False:
    
        for i in range(len(word_incident)):
        
            if word_incident[i] in title:
                incident_list.append([sites,title,url])
                flag = True
                break
        else:
            pass

    if flag == False:
    
        other_list.append([sites,title,url])


def main(request):
    
    list.clear()
    incident_list.clear()
    defense_list.clear()
    other_list.clear()
    overseas_list.clear()
    
    #IPA（重要セキュリテイー情報）のスクレイピング
    url = "https://www.ipa.go.jp/"
    sites = "IPA"
    response = req.get(url)
    soup = bs4(response.content,"html.parser")
    topic = soup.find(class_="top-info --security").find(class_="news-list")
    #topic = soup.find(class_="top-info --security")
    tag_topic = "p"
    tag_url ="a"
    url = url[:-1]
    list_appends_separation(tag_topic,tag_url,sites,topic,url)
    
    #SecurityNextのスクレイピング
    url = "https://www.security-next.com/"
    sites = "Security-Next"
    response = req.get(url)
    soup = bs4(response.content,"html.parser")
    topic = soup.find("dl")
    url = url[:-1]
    list_appends(sites,topic,url)
    
    #ZDNetのスクレイピング
    url = "https://japan.zdnet.com/security/"
    sites = "ZDNet"
    response = req.get(url)
    soup = bs4(response.content,"html.parser")
    topic = soup.find(class_="list-thumb-l")
    url = url[:-1]
    list_appends(sites,topic,url)


    #ITmediaのスクレピング
    url = "https://www.itmedia.co.jp/news/subtop/security/"
    sites = "ITmedia"
    response = req.get(url)
    soup = bs4(response.content,"html.parser")
    topic = soup.find(id="Newarticles")
    url = url[:-1]
    list_appends(sites,topic,url)
     
    
    #海外ニュースサイト(morningstarsecurity)
    url = "https://morningstarsecurity.com/news"
    sites = "Morningstarsecurity"
    response = req.get(url)
    soup = bs4(response.content,"html.parser")
    topic = soup.find(class_="cmra-content-links")
    url = url[:-1]
    overseas_list_appends(sites,topic,url)
    
    
    #海外ニュースサイト(infosecurity-magazine)
    url = "https://www.infosecurity-magazine.com/news/"
    sites = "infosecurity-magazine"
    response = req.get(url)
    soup = bs4(response.content,"html.parser")
    topic = soup.find(class_="webpages-list")
    url = url[:-1]
    overseas_list_appends(sites,topic,url)
    

    #スクレイピングしたデータの処理
    for j in range(len(list)):
        list_classification(list[j][0],list[j][1],list[j][2])
        
    print('ck')
        
    context={'incidentlist':incident_list,
             'defenselist':defense_list,
             'otherlist':other_list,
             'overseaslist':overseas_list,
            }
    
    return render(request,"secnews/main.html",context)

#csv形式で保存
def csvdownload_incident(request):
    
    response = HttpResponse(content_type = 'text/csv; charset=Shift-JIS')
    filename =urllib.parse.quote((u'国内セキュリテイーニュース（インシデント・脆弱性）.csv').encode("utf8"))
    response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'{}'.format(filename)
    writer = csv.writer(response)
    for k in range(len(incident_list) - 1):
        writer.writerow([incident_list[k][0],incident_list[k][1],incident_list[k][2]])
    
    return response

def csvdownload_defense(request):
    
    response = HttpResponse(content_type = 'text/csv; charset=Shift-JIS')
    filename =urllib.parse.quote((u'国内セキュリテイーニュース(対策）.csv').encode("utf8"))
    response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'{}'.format(filename)
    writer = csv.writer(response)
    for k in range(len(defense_list) - 1):
        writer.writerow([defense_list[k][0],defense_list[k][1],defense_list[k][2]])
        
    return response

def csvdownload_other(request):
    
    response = HttpResponse(content_type = 'text/csv; charset=Shift-JIS')
    filename =urllib.parse.quote((u'国内セキュリテイーニュース（その他）.csv').encode("utf8"))
    response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'{}'.format(filename)
    writer = csv.writer(response)
    for k in range(len(other_list) - 1):
        writer.writerow([other_list[k][0],other_list[k][1],other_list[k][2]])
        
    return response
                         
#xlsx形式で出力                        
def xlsxdownload_incident(request):
    
    header = ["サイト名","タイトル","URL"]
    wb = px.Workbook()
    ws = wb.active
    for i in range(len(header)):
        ws.cell(row = 1,column= i+1).value = header[i]
    for i in range(len(incident_list)-1):
            for j in range(len(header)):
                ws.cell(row = 2 + i,column = j+1).value = incident_list[i][j]
    virtual = io.BytesIO()
    wb.save(virtual)
    
    binary = io.BytesIO(virtual.getvalue())
                         
    return FileResponse(binary,filename="国内セキュリテイーニュース（インシデント・脆弱性）.xlsx")


def xlsxdownload_defense(request):
        
    header = ["サイト名","タイトル","URL"]
    wb = px.Workbook()
    ws = wb.active
    for i in range(len(header)):
        ws.cell(row = 1,column= i+1).value = header[i]
    for i in range(len(defense_list)):
            for j in range(len(header)):
                ws.cell(row = 2 + i,column = j+1).value = defense_list[i][j]
    virtual = io.BytesIO()
    wb.save(virtual)
    
    binary = io.BytesIO(virtual.getvalue())
                         
    return FileResponse(binary,filename="国内セキュリテイーニュース（対策）.xlsx")

def xlsxdownload_other(request):
    
    header = ["サイト名","タイトル","URL"]
    wb = px.Workbook()
    ws = wb.active
    for i in range(len(header)):
        ws.cell(row = 1,column= i+1).value = header[i]
    for i in range(len(other_list)):
            for j in range(len(header)):
                ws.cell(row = 2 + i,column = j+1).value = other_list[i][j]
    virtual = io.BytesIO()
    wb.save(virtual)
    
    binary = io.BytesIO(virtual.getvalue())
                         
    return FileResponse(binary,filename="国内セキュリテイーニュース（その他）.xlsx")

#海外のセキュリティニュースの出力
#csv形式
def csvdownload_overseas(request):
    
    response = HttpResponse(content_type = 'text/csv; charset=Shift-JIS')
    filename =urllib.parse.quote((u'海外セキュリティニュース.csv').encode("utf8"))
    response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'{}'.format(filename)
    writer = csv.writer(response)
    for k in range(len(overseas_list) - 1):
        writer.writerow([overseas_list[k][0],overseas_list[k][1],overseas_list[k][2]])
        
    return response
                         
#xlsx形式で出力                        
def xlsxdownload_overseas(request):
    
    header = ["サイト名","タイトル","URL"]
    wb = px.Workbook()
    ws = wb.active
    for i in range(len(header)):
        ws.cell(row = 1,column= i+1).value = header[i]
    for i in range(len(overseas_list)-1):
            for j in range(len(header)):
                ws.cell(row = 2 + i,column = j+1).value = overseas_list[i][j]
    virtual = io.BytesIO()
    wb.save(virtual)
    
    binary = io.BytesIO(virtual.getvalue())
                         
    return FileResponse(binary,filename="海外セキュリテイーニュース.xlsx")
